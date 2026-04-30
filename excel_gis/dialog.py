# -*- coding: utf-8 -*-
"""
LVT4U Excel → GIS — Import Excel coordinates to QGIS vector layer.

Core flow:
    1. Select Excel file (.xlsx)
    2. Select Sheet — preview data in table
    3. Map columns: X, Y, optional Group (plot name), Point Order
    4. Select CRS (VN-2000 meter or WGS84)
    5. Choose output: Points or Polygon (group by plot)
    6. Import → create memory layer on map

Author: Lộc Vũ Trung (LVT) / Slow Forest
License: GPL-3.0
"""

import os

from qgis.PyQt.QtCore import Qt, QVariant
from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QLabel, QComboBox, QPushButton, QFileDialog, QMessageBox,
    QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QRadioButton, QButtonGroup, QProgressBar, QApplication,
    QTabWidget, QWidget, QTextBrowser, QCheckBox,
)
from qgis.PyQt.QtGui import QIcon, QColor
from qgis.core import (
    QgsProject, QgsVectorLayer, QgsFeature, QgsGeometry,
    QgsPointXY, QgsField, QgsFields,
    QgsCoordinateReferenceSystem,
)
from qgis.gui import QgsProjectionSelectionWidget

from ..shared.i18n import current_language

# Bilingual labels
_VI = {
    "Excel → GIS": "Chuyển Excel vào bản đồ",
    "File:": "Tệp:", "Browse...": "Chọn...",
    "Sheet:": "Trang tính:", "Data Preview": "Xem trước dữ liệu",
    "Column Mapping": "Ánh xạ cột",
    "X Column:": "Cột X:", "Y Column:": "Cột Y:",
    "Group Column:": "Cột nhóm (lô):", "Order Column:": "Cột thứ tự:",
    "CRS": "Hệ tọa độ",
    "Output Type": "Loại đầu ra",
    "Points": "Điểm", "Polygon (group by plot)": "Vùng (nhóm theo lô)",
    "Import": "Nhập", "Close": "Đóng", "Guide": "Hướng dẫn",
    "No file selected.": "Chưa chọn tệp.",
    "Select X and Y columns.": "Chọn cột X và Y.",
    "Select Group and Order columns for polygon.":
        "Chọn cột Nhóm và Thứ tự để tạo polygon.",
    "(none)": "(không)",
    "Layer Name:": "Tên lớp:",
    "Save to file": "Lưu ra file",
    "rows": "hàng",
    "Import Data": "Nhập dữ liệu",
    "imported successfully!": "đã nhập thành công!",
    "features": "đối tượng",
}


def _t(text, lang=None):
    if lang is None:
        lang = current_language()
    return _VI.get(text, text) if lang == 'vi' else text


class ExcelGisDialog(QDialog):
    """Dialog for importing Excel coordinate data into QGIS."""

    def __init__(self, iface, parent=None):
        super().__init__(parent or iface.mainWindow())
        self.iface = iface
        self.lang = current_language()
        self._wb = None           # openpyxl Workbook
        self._headers = []        # current sheet headers
        self._data_rows = []      # current sheet data (list of lists)
        self._file_path = ""

        self.setMinimumSize(750, 600)
        self._setup_ui()
        self._refresh_ui_text()

    # ------------------------------------------------------------------
    # UI Construction
    # ------------------------------------------------------------------

    def _setup_ui(self):
        main_ly = QVBoxLayout(self)

        # Language toggle
        top = QHBoxLayout()
        self.btn_lang = QPushButton("🇬🇧 English")
        self.btn_lang.setFixedWidth(120)
        self.btn_lang.clicked.connect(self._toggle_lang)
        top.addWidget(self.btn_lang)
        top.addStretch()
        main_ly.addLayout(top)

        # Tabs
        self.tabs = QTabWidget()
        main_ly.addWidget(self.tabs)

        # Tab 1: Main
        self.tab_main = QWidget()
        self.tabs.addTab(self.tab_main, "Import Data")
        self._build_main_tab()

        # Tab 2: Guide
        self.tab_guide = QWidget()
        self.tabs.addTab(self.tab_guide, "Guide")
        self._build_guide_tab()

    def _build_main_tab(self):
        ly = QVBoxLayout(self.tab_main)

        # --- File Selection ---
        file_row = QHBoxLayout()
        self.lbl_file = QLabel("File:")
        self.lbl_file_path = QLabel("—")
        self.lbl_file_path.setStyleSheet("color: #555;")
        self.btn_browse = QPushButton("Browse...")
        self.btn_browse.clicked.connect(self._browse_file)
        file_row.addWidget(self.lbl_file)
        file_row.addWidget(self.lbl_file_path, 1)
        file_row.addWidget(self.btn_browse)
        ly.addLayout(file_row)

        # --- Sheet Selection ---
        sheet_row = QHBoxLayout()
        self.lbl_sheet = QLabel("Sheet:")
        self.cmb_sheet = QComboBox()
        self.cmb_sheet.currentIndexChanged.connect(self._on_sheet_changed)
        sheet_row.addWidget(self.lbl_sheet)
        sheet_row.addWidget(self.cmb_sheet, 1)
        self.lbl_rows = QLabel("")
        sheet_row.addWidget(self.lbl_rows)
        ly.addLayout(sheet_row)

        # --- Data Preview ---
        self.grp_preview = QGroupBox("Data Preview")
        pv_ly = QVBoxLayout()
        self.tbl_preview = QTableWidget()
        self.tbl_preview.setMaximumHeight(200)
        self.tbl_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        pv_ly.addWidget(self.tbl_preview)
        self.grp_preview.setLayout(pv_ly)
        ly.addWidget(self.grp_preview)

        # --- Column Mapping + CRS (side by side) ---
        mid_row = QHBoxLayout()

        # Left: Column mapping
        self.grp_cols = QGroupBox("Column Mapping")
        col_ly = QFormLayout()
        self.cmb_x = QComboBox()
        self.cmb_y = QComboBox()
        self.cmb_group = QComboBox()
        self.cmb_order = QComboBox()
        col_ly.addRow("X Column:", self.cmb_x)
        col_ly.addRow("Y Column:", self.cmb_y)
        col_ly.addRow("Group Column:", self.cmb_group)
        col_ly.addRow("Order Column:", self.cmb_order)
        self.grp_cols.setLayout(col_ly)
        mid_row.addWidget(self.grp_cols, 1)

        # Right: CRS + Output Type
        right_ly = QVBoxLayout()

        self.grp_crs = QGroupBox("CRS")
        crs_ly = QVBoxLayout()
        self.crs_widget = QgsProjectionSelectionWidget()
        self.crs_widget.setCrs(
            QgsCoordinateReferenceSystem("EPSG:4326")
        )
        crs_ly.addWidget(self.crs_widget)
        self.grp_crs.setLayout(crs_ly)
        right_ly.addWidget(self.grp_crs)

        self.grp_output = QGroupBox("Output Type")
        out_ly = QVBoxLayout()
        self.rad_points = QRadioButton("Points")
        self.rad_polygon = QRadioButton("Polygon (group by plot)")
        self.rad_points.setChecked(True)
        self.rad_polygon.toggled.connect(self._on_output_type_changed)
        out_ly.addWidget(self.rad_points)
        out_ly.addWidget(self.rad_polygon)
        self.grp_output.setLayout(out_ly)
        right_ly.addWidget(self.grp_output)

        mid_row.addLayout(right_ly, 1)
        ly.addLayout(mid_row)

        # --- Layer name ---
        name_row = QHBoxLayout()
        self.lbl_layer_name = QLabel("Layer Name:")
        self.txt_layer_name = QComboBox()
        self.txt_layer_name.setEditable(True)
        self.txt_layer_name.setCurrentText("Excel_Import")
        name_row.addWidget(self.lbl_layer_name)
        name_row.addWidget(self.txt_layer_name, 1)

        self.chk_save = QCheckBox("Save to file")
        name_row.addWidget(self.chk_save)
        ly.addLayout(name_row)

        # --- Progress + Buttons ---
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        ly.addWidget(self.progress)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_import = QPushButton("⬇️ Import")
        self.btn_import.setFixedHeight(36)
        self.btn_import.setStyleSheet(
            "QPushButton { background: #2e7d32; color: white; "
            "font-weight: bold; padding: 4px 20px; border-radius: 4px; }"
            "QPushButton:hover { background: #388e3c; }"
        )
        self.btn_import.clicked.connect(self._do_import)
        btn_row.addWidget(self.btn_import)
        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.close)
        btn_row.addWidget(self.btn_close)
        ly.addLayout(btn_row)

    def _build_guide_tab(self):
        ly = QVBoxLayout(self.tab_guide)
        self.guide_browser = QTextBrowser()
        self.guide_browser.setOpenExternalLinks(True)
        ly.addWidget(self.guide_browser)
        self._set_guide_content()

    def _set_guide_content(self):
        if self.lang == 'vi':
            html = """
            <h2>Hướng dẫn — Chuyển Excel vào bản đồ</h2>
            <h3>Bước 1 — Chọn file Excel</h3>
            <p>Nhấn <b>Chọn...</b> để chọn file Excel (.xlsx).
            Hệ thống sẽ tự đọc các trang tính (sheet).</p>
            <h3>Bước 2 — Xem trước dữ liệu</h3>
            <p>Chọn sheet chứa dữ liệu tọa độ. Bảng preview hiển thị
            20 dòng đầu tiên.</p>
            <h3>Bước 3 — Ánh xạ cột</h3>
            <p>Chọn cột chứa <b>X</b> (kinh độ/Easting) và <b>Y</b>
            (vĩ độ/Northing).</p>
            <p>Nếu muốn tạo <b>Polygon</b>: chọn thêm cột Nhóm (tên lô)
            và cột Thứ tự điểm.</p>
            <h3>Bước 4 — Hệ tọa độ</h3>
            <p>Chọn hệ tọa độ phù hợp:</p>
            <ul>
            <li><b>EPSG:4326</b> — WGS84 (độ/phút/giây)</li>
            <li><b>VN-2000</b> — Hệ mét (tìm theo tên tỉnh)</li>
            </ul>
            <h3>Bước 5 — Nhập</h3>
            <p>Nhấn <b>Nhập</b> → Layer mới được tạo trên bản đồ QGIS.</p>
            """
        else:
            html = """
            <h2>Guide — Excel → GIS</h2>
            <h3>Step 1 — Select Excel file</h3>
            <p>Click <b>Browse...</b> to select an Excel file (.xlsx).
            Sheets will be loaded automatically.</p>
            <h3>Step 2 — Preview data</h3>
            <p>Select the sheet with coordinate data. Preview shows
            the first 20 rows.</p>
            <h3>Step 3 — Map columns</h3>
            <p>Select columns for <b>X</b> (Longitude/Easting) and <b>Y</b>
            (Latitude/Northing).</p>
            <p>For <b>Polygon</b> output: also select Group (plot name)
            and Point Order columns.</p>
            <h3>Step 4 — CRS</h3>
            <p>Select the coordinate reference system:</p>
            <ul>
            <li><b>EPSG:4326</b> — WGS84 (degrees)</li>
            <li><b>VN-2000</b> — Metric (search by province name)</li>
            </ul>
            <h3>Step 5 — Import</h3>
            <p>Click <b>Import</b> → A new layer is added to QGIS map.</p>
            """
        self.guide_browser.setHtml(html)

    # ------------------------------------------------------------------
    # Language
    # ------------------------------------------------------------------

    def _toggle_lang(self):
        self.lang = 'en' if self.lang == 'vi' else 'vi'
        self._refresh_ui_text()

    def _refresh_ui_text(self):
        t = lambda s: _t(s, self.lang)
        self.setWindowTitle(f"LVT4U — {t('Excel → GIS')}")
        self.btn_lang.setText(
            "🇬🇧 English" if self.lang == 'en' else "🇻🇳 Tiếng Việt"
        )
        self.tabs.setTabText(0, t("Import Data"))
        self.tabs.setTabText(1, t("Guide"))
        self.lbl_file.setText(t("File:"))
        self.btn_browse.setText(t("Browse..."))
        self.lbl_sheet.setText(t("Sheet:"))
        self.grp_preview.setTitle(t("Data Preview"))
        self.grp_cols.setTitle(t("Column Mapping"))
        self.grp_crs.setTitle(t("CRS"))
        self.grp_output.setTitle(t("Output Type"))
        self.rad_points.setText(t("Points"))
        self.rad_polygon.setText(t("Polygon (group by plot)"))
        self.lbl_layer_name.setText(t("Layer Name:"))
        self.chk_save.setText(t("Save to file"))
        self.btn_import.setText(f"⬇️ {t('Import')}")
        self.btn_close.setText(t("Close"))

        # Update form labels in column mapping
        form = self.grp_cols.layout()
        if form:
            labels = [t("X Column:"), t("Y Column:"),
                      t("Group Column:"), t("Order Column:")]
            for i, lbl_text in enumerate(labels):
                item = form.itemAt(i, QFormLayout.LabelRole)
                if item and item.widget():
                    item.widget().setText(lbl_text)

        self._set_guide_content()

    # ------------------------------------------------------------------
    # File / Sheet
    # ------------------------------------------------------------------

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, _t("Select Excel File", self.lang), "",
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if not path:
            return
        self._file_path = path
        self.lbl_file_path.setText(os.path.basename(path))
        self.lbl_file_path.setToolTip(path)
        self._load_workbook(path)

    def _load_workbook(self, path):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self, "LVT4U",
                "openpyxl is required.\n"
                "Install: pip install openpyxl"
            )
            return

        try:
            self._wb = openpyxl.load_workbook(path, read_only=True,
                                               data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "LVT4U", f"Error reading file:\n{e}")
            return

        self.cmb_sheet.blockSignals(True)
        self.cmb_sheet.clear()
        for name in self._wb.sheetnames:
            self.cmb_sheet.addItem(name)
        self.cmb_sheet.blockSignals(False)

        # Auto-set layer name from filename
        base = os.path.splitext(os.path.basename(path))[0]
        self.txt_layer_name.setCurrentText(base)

        self._on_sheet_changed()

    def _on_sheet_changed(self):
        if not self._wb:
            return
        sheet_name = self.cmb_sheet.currentText()
        if not sheet_name:
            return

        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # First row = headers — deduplicate names
        raw_headers = [str(h).strip() if h else f"Col_{i}"
                       for i, h in enumerate(rows[0])]
        seen = {}
        deduped = []
        for h in raw_headers:
            if h in seen:
                seen[h] += 1
                deduped.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                deduped.append(h)
        self._headers = deduped
        self._data_rows = rows[1:]

        self.lbl_rows.setText(
            f"{len(self._data_rows)} {_t('rows', self.lang)}"
        )

        # Update preview table (max 20 rows)
        preview = self._data_rows[:20]
        self.tbl_preview.setColumnCount(len(self._headers))
        self.tbl_preview.setRowCount(len(preview))
        self.tbl_preview.setHorizontalHeaderLabels(self._headers)

        for r, row in enumerate(preview):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                self.tbl_preview.setItem(r, c, item)

        self.tbl_preview.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents
        )

        # Update column combos
        self._populate_column_combos()

    def _populate_column_combos(self):
        none_text = _t("(none)", self.lang)

        for cmb in (self.cmb_x, self.cmb_y, self.cmb_group, self.cmb_order):
            cmb.blockSignals(True)
            cmb.clear()

        # X and Y don't have "(none)"
        for h in self._headers:
            self.cmb_x.addItem(h)
            self.cmb_y.addItem(h)

        # Group and Order have optional "(none)"
        self.cmb_group.addItem(none_text, None)
        self.cmb_order.addItem(none_text, None)
        for h in self._headers:
            self.cmb_group.addItem(h, h)
            self.cmb_order.addItem(h, h)

        # Auto-detect X/Y columns
        x_keys = ['x', 'lon', 'longitude', 'easting', 'kinh_do', 'kinhdo',
                   'long', 'e', 'east']
        y_keys = ['y', 'lat', 'latitude', 'northing', 'vi_do', 'vido',
                   'north', 'n']
        group_keys = ['ten_lo', 'tenlo', 'lo', 'plot', 'group', 'name',
                      'ten', 'ma_lo', 'malo']
        order_keys = ['thu_tu', 'thutu', 'stt', 'order', 'point_order',
                      'tt', 'seq', 'point_no']

        for i, h in enumerate(self._headers):
            hl = h.lower().strip()
            if hl in x_keys:
                self.cmb_x.setCurrentIndex(i)
            if hl in y_keys:
                self.cmb_y.setCurrentIndex(i)
            # +1 offset because group/order have "(none)" at index 0
            if hl in group_keys:
                self.cmb_group.setCurrentIndex(i + 1)
            if hl in order_keys:
                self.cmb_order.setCurrentIndex(i + 1)

        for cmb in (self.cmb_x, self.cmb_y, self.cmb_group, self.cmb_order):
            cmb.blockSignals(False)

    def _on_output_type_changed(self):
        is_poly = self.rad_polygon.isChecked()
        self.cmb_group.setEnabled(is_poly)
        self.cmb_order.setEnabled(is_poly)

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    def _do_import(self):
        if not self._data_rows:
            QMessageBox.warning(
                self, "LVT4U", _t("No file selected.", self.lang)
            )
            return

        x_col = self.cmb_x.currentText()
        y_col = self.cmb_y.currentText()
        if not x_col or not y_col:
            QMessageBox.warning(
                self, "LVT4U", _t("Select X and Y columns.", self.lang)
            )
            return

        is_polygon = self.rad_polygon.isChecked()
        group_col = self.cmb_group.currentData()
        order_col = self.cmb_order.currentData()

        if is_polygon and not group_col:
            QMessageBox.warning(
                self, "LVT4U",
                _t("Select Group and Order columns for polygon.", self.lang)
            )
            return

        x_idx = self._headers.index(x_col)
        y_idx = self._headers.index(y_col)
        group_idx = self._headers.index(group_col) if group_col else None
        order_idx = self._headers.index(order_col) if order_col else None

        crs = self.crs_widget.crs()
        layer_name = self.txt_layer_name.currentText() or "Excel_Import"

        self.progress.setVisible(True)
        self.progress.setValue(0)
        QApplication.processEvents()

        try:
            if is_polygon:
                count = self._import_polygons(
                    x_idx, y_idx, group_idx, order_idx,
                    crs, layer_name
                )
            else:
                count = self._import_points(
                    x_idx, y_idx, crs, layer_name
                )
        except Exception as e:
            self.progress.setVisible(False)
            QMessageBox.critical(self, "LVT4U", f"Import error:\n{e}")
            return

        self.progress.setValue(100)
        QApplication.processEvents()

        QMessageBox.information(
            self, "LVT4U",
            f"{count} {_t('features', self.lang)} "
            f"{_t('imported successfully!', self.lang)}"
        )
        self.progress.setVisible(False)

    def _import_points(self, x_idx, y_idx, crs, layer_name):
        """Import rows as point features."""
        uri = f"Point?crs={crs.authid()}"
        layer = QgsVectorLayer(uri, layer_name, "memory")
        pr = layer.dataProvider()

        # Add all columns as fields
        fields = QgsFields()
        for h in self._headers:
            fields.append(QgsField(h, QVariant.String))
        pr.addAttributes(fields)
        layer.updateFields()

        features = []
        total = len(self._data_rows)
        for i, row in enumerate(self._data_rows):
            if i % 100 == 0:
                self.progress.setValue(int(i / total * 90))
                QApplication.processEvents()

            try:
                x = float(row[x_idx])
                y = float(row[y_idx])
            except (TypeError, ValueError, IndexError):
                continue

            feat = QgsFeature(layer.fields())
            feat.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(x, y)))

            for j, h in enumerate(self._headers):
                val = row[j] if j < len(row) else None
                feat.setAttribute(h, str(val) if val is not None else "")

            features.append(feat)

        pr.addFeatures(features)
        layer.updateExtents()
        QgsProject.instance().addMapLayer(layer)
        self.iface.mapCanvas().setExtent(layer.extent())
        self.iface.mapCanvas().refresh()
        return len(features)

    def _import_polygons(self, x_idx, y_idx, group_idx, order_idx,
                         crs, layer_name):
        """Import rows grouped into polygons."""
        # Group rows by plot name
        groups = {}
        for row in self._data_rows:
            try:
                gname = str(row[group_idx]) if row[group_idx] else "Unknown"
                x = float(row[x_idx])
                y = float(row[y_idx])
                order = 0
                if order_idx is not None:
                    try:
                        order = float(row[order_idx])
                    except (TypeError, ValueError):
                        order = 0
            except (TypeError, ValueError, IndexError):
                continue

            if gname not in groups:
                groups[gname] = []
            groups[gname].append((order, x, y, row))

        # Sort each group by order
        for gname in groups:
            groups[gname].sort(key=lambda t: t[0])

        uri = f"Polygon?crs={crs.authid()}"
        layer = QgsVectorLayer(uri, layer_name, "memory")
        pr = layer.dataProvider()

        # Fields: group name + extra columns (excluding X/Y/Group/Order)
        fields = QgsFields()
        fields.append(QgsField("plot_name", QVariant.String))
        fields.append(QgsField("point_count", QVariant.Int))
        pr.addAttributes(fields)
        layer.updateFields()

        features = []
        total = len(groups)
        for i, (gname, points) in enumerate(groups.items()):
            self.progress.setValue(int(i / total * 90))
            QApplication.processEvents()

            if len(points) < 3:
                continue  # Need at least 3 points for polygon

            ring = [QgsPointXY(p[1], p[2]) for p in points]
            # Close the ring
            ring.append(ring[0])

            feat = QgsFeature(layer.fields())
            feat.setGeometry(QgsGeometry.fromPolygonXY([ring]))
            feat.setAttribute("plot_name", gname)
            feat.setAttribute("point_count", len(points))
            features.append(feat)

        pr.addFeatures(features)
        layer.updateExtents()
        QgsProject.instance().addMapLayer(layer)
        self.iface.mapCanvas().setExtent(layer.extent())
        self.iface.mapCanvas().refresh()
        return len(features)

    # ------------------------------------------------------------------
    def refresh_layers(self):
        """Called by plugin coordinator on show."""
        pass
